import streamlit as st
import streamlit.components.v1 as components
from pathlib import Path
from cable_list_component import cable_list_widget
from pdf_viewer_component import pdf_viewer_widget
import fitz  # PyMuPDF
import re
import io
import base64
import math
import datetime
import string
import copy
import zipfile
import json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font as XLFont
from openpyxl.utils import get_column_letter

st.set_page_config(
    page_title="CablingToolPro",
    page_icon="🔌",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────────────────────────────────────
# Helpers (port from desktop app)
# ─────────────────────────────────────────────────────────────────────────────

def _natural_key(s):
    """Key for natural (human) sort: splits text into str/int chunks."""
    return [int(c) if c.isdigit() else c.lower() for c in re.split(r'(\d+)', s)]


def kabel_label_alpha(idx):
    group_size = 24
    letter = chr(ord("A") + (idx // group_size))
    number = (idx % group_size) + 1
    return f"{letter}.{number:02d}"


def staffelpreis(treffer):
    staffeln = [
        (50, 2.20), (50, 2.05), (50, 1.90), (50, 1.80), (50, 1.70),
        (50, 1.60), (50, 1.50), (50, 1.40), (50, 1.30), (50, 1.20),
        (50, 1.10), (50, 1.05), (50, 1.00), (50, 0.98), (50, 0.96),
        (50, 0.94), (50, 0.92), (50, 0.90), (50, 0.88), (50, 0.86),
        (50, 0.84), (50, 0.83), (50, 0.82), (50, 0.81), (50, 0.81),
        (50, 0.81), (50, 0.81), (50, 0.81), (50, 0.81), (50, 0.81),
        (50, 0.81), (50, 0.81), (50, 0.81), (50, 0.81), (50, 0.81),
        (50, 0.81), (50, 0.81), (50, 0.81), (50, 0.80),
    ]
    preis = 0
    rest = treffer
    for stufenmenge, stufenpreis_val in staffeln:
        if rest > stufenmenge:
            preis += stufenmenge * stufenpreis_val
            rest -= stufenmenge
        else:
            preis += rest * stufenpreis_val
            rest = 0
            break
    if rest > 0:
        preis += rest * 0.75
    return preis


LABEL_BG_COLORS_HEX = ["#ffe066", "#7eeaff", "#a7ef7e", "#ff8cdd", "#c89fff"]


def hex_to_fitz_color(hex_color):
    h = hex_color.lstrip("#")
    return (int(h[0:2], 16) / 255.0, int(h[2:4], 16) / 255.0, int(h[4:6], 16) / 255.0)


# ─────────────────────────────────────────────────────────────────────────────
# Session state initialisation
# ─────────────────────────────────────────────────────────────────────────────

def _init():
    defaults = {
        "doc_bytes": None,
        "kabel_fields": [],
        "annotations": [],
        "search_terms": [],
        "current_page": 0,
        "zoom": 1.5,
        "setting_2x_rj45": True,
        "setting_2x_short": True,
        "setting_rj45": True,
        "setting_2xukv": False,
        "export_pdf_bytes": None,
        "search_ran": False,
        "_id_seq": 0,
        "kabel_fields_snap": [],
        "annotations_snap": [],
        "pdf_dirty": False,
        "user_name": "",
        "user_email": "",
        "user_registered": False,
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v


_init()


def _new_id():
    """Return a session-persistent, ever-increasing stable ID for kabel_fields entries."""
    st.session_state["_id_seq"] += 1
    return st.session_state["_id_seq"]


def send_invoice_email(user_name, user_email, active_count, preis, excel_bytes, filename):
    """Send an invoice e-mail to the admin with the Excel file attached.
    Credentials are read from st.secrets (never hard-coded).
    Silently skips if secrets are not configured."""
    import smtplib
    import ssl
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders

    try:
        cfg = st.secrets.get("email", {})
        smtp_host     = cfg.get("smtp_host", "")
        smtp_port     = int(cfg.get("smtp_port", 587))
        smtp_user     = cfg.get("smtp_user", "")
        smtp_password = cfg.get("smtp_password", "")
        admin_email   = cfg.get("admin_email", smtp_user)
        if not smtp_host or not smtp_user or not smtp_password:
            return  # secrets not configured — skip silently

        msg = MIMEMultipart()
        msg["From"]    = smtp_user
        msg["To"]      = admin_email
        msg["Subject"] = f"CablingToolPro – Export von {user_name}"

        body = (
            f"Neuer Export\n\n"
            f"Name:          {user_name}\n"
            f"E-Mail:        {user_email}\n"
            f"Beschriftungen:{active_count}\n"
            f"Betrag:        CHF {preis:.2f}\n"
            f"Datei:         {filename}\n"
            f"Datum:         {datetime.date.today()}\n"
        )
        msg.attach(MIMEText(body, "plain"))

        part = MIMEBase("application", "octet-stream")
        part.set_payload(excel_bytes)
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f'attachment; filename="{filename}"')
        msg.attach(part)

        context = ssl.create_default_context()
        with smtplib.SMTP(smtp_host, smtp_port) as server:
            server.ehlo()
            server.starttls(context=context)
            server.login(smtp_user, smtp_password)
            server.sendmail(smtp_user, admin_email, msg.as_string())
    except Exception:
        pass  # never crash the app over a notification email


def _clear_label_widgets():
    """Delete stale text_input / checkbox widget states so they re-initialize."""
    for key in list(st.session_state.keys()):
        if key.startswith("label_") or key.startswith("cb_"):
            del st.session_state[key]


def _build_project_zip():
    """Pack current session state into a .ctpro ZIP (PDF + JSON metadata)."""
    meta = {
        "kabel_fields":      st.session_state.kabel_fields,
        "kabel_fields_snap": st.session_state.kabel_fields_snap,
        "annotations":       st.session_state.annotations,
        "annotations_snap":  st.session_state.annotations_snap,
        "search_terms":      st.session_state.search_terms,
        "settings": {
            "setting_2x_rj45":  st.session_state.setting_2x_rj45,
            "setting_2x_short": st.session_state.setting_2x_short,
            "setting_rj45":     st.session_state.setting_rj45,
            "setting_2xukv":    st.session_state.setting_2xukv,
        },
        "_id_seq": st.session_state["_id_seq"],
    }
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr("plan.pdf", st.session_state.doc_bytes)
        zf.writestr("project.json", json.dumps(meta, ensure_ascii=False))
    buf.seek(0)
    return buf.getvalue()


def _load_project(zip_bytes):
    """Restore session state from a .ctpro ZIP file."""
    with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
        pdf_bytes = zf.read("plan.pdf")
        meta      = json.loads(zf.read("project.json").decode("utf-8"))
    st.session_state.doc_bytes          = pdf_bytes
    st.session_state.kabel_fields       = meta.get("kabel_fields", [])
    st.session_state.kabel_fields_snap  = meta.get("kabel_fields_snap", copy.deepcopy(meta.get("kabel_fields", [])))
    st.session_state.annotations        = meta.get("annotations", [])
    st.session_state.annotations_snap   = meta.get("annotations_snap", list(meta.get("annotations", [])))
    st.session_state.search_terms       = meta.get("search_terms", [])
    settings = meta.get("settings", {})
    st.session_state.setting_2x_rj45  = settings.get("setting_2x_rj45",  True)
    st.session_state.setting_2x_short = settings.get("setting_2x_short", True)
    st.session_state.setting_rj45     = settings.get("setting_rj45",     True)
    st.session_state.setting_2xukv    = settings.get("setting_2xukv",    False)
    # Restore ID counter so new IDs never collide with loaded ones
    existing_ids = [k.get("_id", 0) for k in st.session_state.kabel_fields]
    st.session_state["_id_seq"] = max(existing_ids + [meta.get("_id_seq", 0)])
    st.session_state.pdf_dirty        = False
    st.session_state.export_pdf_bytes = None
    st.session_state.search_ran       = bool(st.session_state.kabel_fields)
    kf = st.session_state.kabel_fields
    st.session_state.current_page = kf[0]["page_num"] if kf else 0
    _clear_label_widgets()
    _clear_component_states()


def _clear_component_states():
    """Clear cached component values so stale data never overwrites kabel_fields."""
    for key in list(st.session_state.keys()):
        if key.startswith("cl_"):
            del st.session_state[key]


# ─────────────────────────────────────────────────────────────────────────────
# Core search logic (ported from AcrobatViewer.search_and_highlight)
# ─────────────────────────────────────────────────────────────────────────────

def search_pdf(doc_bytes, terms, s2x, s2x_short, s1x, s2xukv):
    pat2 = [r"2x\s*rj\s*45", r"2\s*x\s*rj\s*45", r"2xrj45", r"2 x rj45"]
    pat2_short = [r"\b2x\b"]
    pat1 = [r"rj\s*45", r"rj45"]
    pat2xukv = [r"2xukv"]

    doc = fitz.open(stream=doc_bytes, filetype="pdf")
    ukv_hits = {t.lower(): [] for t in terms}
    rj_hits_dict = {}

    for pnum in range(len(doc)):
        pg = doc.load_page(pnum)
        dat = pg.get_text("dict")

        lines = []
        for blk in dat["blocks"]:
            for ln in blk.get("lines", []):
                for span in ln.get("spans", []):
                    txt = span["text"].strip()
                    bb = fitz.Rect(span["bbox"])
                    c = span.get("color", 0)
                    hex_c = "#%02x%02x%02x" % ((c >> 16) & 0xFF, (c >> 8) & 0xFF, c & 0xFF)
                    lines.append((txt, bb, hex_c))

        rj_hits = []
        for txt, bb, _ in lines:
            norm = re.sub(r"\s+", "", txt.lower())
            if s2x and any(re.search(p, norm) for p in pat2):
                rj_hits.append({"type": "2x RJ45", "rect": bb, "used": False, "kabel_typ": "2x RJ45"})
            elif s2x_short and any(re.search(p, txt.lower()) for p in pat2_short):
                rj_hits.append({"type": "2x RJ45", "rect": bb, "used": False, "kabel_typ": "2x RJ45"})
            elif s2xukv and any(re.search(p, norm) for p in pat2xukv):
                rj_hits.append({"type": "2x UKV", "rect": bb, "used": False, "kabel_typ": "2x RJ45"})
            elif (
                s1x
                and any(re.search(p, norm) for p in pat1)
                and not any(re.search(p, norm) for p in pat2)
                and not (s2xukv and any(re.search(p, norm) for p in pat2xukv))
            ):
                rj_hits.append({"type": "RJ45", "rect": bb, "used": False, "kabel_typ": "RJ45"})
        rj_hits_dict[pnum] = rj_hits

        for txt, bb, color_hex in lines:
            for term in terms:
                if term.lower() in txt.lower():
                    ukv_hits[term.lower()].append({
                        "txt": txt, "bb": bb, "page_num": pnum,
                        "is_combined": False, "color_hex": color_hex,
                    })

        # 2-line combined search
        i = 0
        while i < len(lines) - 1:
            txt1, bb1, hex1 = lines[i]
            txt2, bb2, _ = lines[i + 1]
            combined = f"{txt1} {txt2}"
            found = False
            for term in terms:
                if (
                    term.lower() in combined.lower()
                    and term.lower() not in txt1.lower()
                    and term.lower() not in txt2.lower()
                ):
                    union = fitz.Rect(
                        min(bb1.x0, bb2.x0), min(bb1.y0, bb2.y0),
                        max(bb1.x1, bb2.x1), max(bb1.y1, bb2.y1),
                    )
                    ukv_hits[term.lower()].append({
                        "txt": combined, "bb": union, "page_num": pnum,
                        "is_combined": True, "color_hex": hex1,
                    })
                    found = True
                    break
            i += 2 if found else 1

    doc.close()

    kabel_fields = []
    annotations = []

    for idx_t, term in enumerate(terms):
        term_lower = term.lower()
        hits = sorted(ukv_hits.get(term_lower, []), key=lambda h: (h["bb"].x0, h["bb"].y0))
        bg_hex = LABEL_BG_COLORS_HEX[idx_t % len(LABEL_BG_COLORS_HEX)]
        cnt = 0

        for hit in hits:
            pnum = hit["page_num"]
            bb = hit["bb"]

            rj_hits = rj_hits_dict[pnum]
            closest, md = None, float("inf")
            for rj in rj_hits:
                if rj["used"]:
                    continue
                d = math.hypot(rj["rect"].x0 - bb.x0, rj["rect"].y0 - bb.y0)
                if d < 70 and d < md:
                    closest, md = rj, d

            if closest:
                if "2x" in closest["type"]:
                    cable_label = f"{kabel_label_alpha(cnt)}/{kabel_label_alpha(cnt + 1)}"
                    kabel_typ = "2x RJ45"
                    col = "green"
                    cnt += 2
                else:
                    cable_label = kabel_label_alpha(cnt)
                    kabel_typ = "RJ45"
                    col = "blue"
                    cnt += 1
                annotations.append({
                    "page_num": pnum,
                    "rect": [closest["rect"].x0, closest["rect"].y0,
                             closest["rect"].x1, closest["rect"].y1],
                    "color": col,
                })
                closest["used"] = True
            else:
                cable_label = kabel_label_alpha(cnt)
                kabel_typ = None
                cnt += 1

            annotations.append({
                "page_num": pnum,
                "rect": [bb.x0, bb.y0, bb.x1, bb.y1],
                "color": "yellow",
            })

            kabel_fields.append({
                "page_num": pnum,
                "rect": [bb.x0, bb.y0, bb.x1, bb.y1],
                "ukv_text": hit["txt"],
                "term": term,
                "label": cable_label,
                "kabel_typ": kabel_typ,
                "label_bg_color": bg_hex,
                "checked": True,
                "_id": _new_id(),
            })

    return kabel_fields, annotations


# ─────────────────────────────────────────────────────────────────────────────
# PDF rendering
# ─────────────────────────────────────────────────────────────────────────────

def _rects_overlap(ax0, ay0, ax1, ay1, bx0, by0, bx1, by1, gap=1):
    """True if the two rectangles overlap or are closer than `gap` pts."""
    return ax0 < bx1 + gap and ax1 > bx0 - gap and ay0 < by1 + gap and ay1 > by0 - gap


def _place_label(occupied, x0, y0, box_w, box_h):
    """Return (lx0, ly0) that doesn't overlap any rect in `occupied`.
    Tries directly above the anchor first, then shifts right, then downward rows.
    All positions are plain Python floats — no fitz.Rect mutation."""
    # Try starting position: directly above anchor point
    lx, ly = x0, y0 - box_h - 1
    for _ in range(400):
        clash = False
        for (ox0, oy0, ox1, oy1) in occupied:
            if _rects_overlap(lx, ly, lx + box_w, ly + box_h, ox0, oy0, ox1, oy1):
                lx = ox1 + 2          # shift right past the clashing box
                clash = True
                break
        if not clash:
            return lx, ly
    # Fallback: place one row higher
    return x0, y0 - box_h * 2 - 2


def _draw_label_annot(page, x0, y0, box_w, box_h, fill_rgb, text, fs, pad_h):
    """Draw background rect_annot + centered freetext_annot on `page`."""
    bg_ann = page.add_rect_annot(fitz.Rect(x0, y0, x0 + box_w, y0 + box_h))
    bg_ann.set_colors(fill=fill_rgb, stroke=fill_rgb)
    bg_ann.set_border(width=0)
    bg_ann.update()
    bg_center = y0 + box_h / 2
    txt_y0 = bg_center - fs / 2
    txt_y1 = bg_center + fs / 2
    txt_ann = page.add_freetext_annot(
        fitz.Rect(x0 + pad_h, txt_y0, x0 + box_w - pad_h, txt_y1),
        text, fontsize=fs, fontname="Helv",
        text_color=(0, 0, 0), fill_color=None,
        rotate=0, align=1,
    )
    txt_ann.set_border(width=0)
    txt_ann.update()


def render_page(doc_bytes, page_num, kabel_fields, annotations, zoom=1.5):
    doc = fitz.open(stream=doc_bytes, filetype="pdf")
    page = doc.load_page(page_num)

    stroke_map = {
        "green": (0, 1, 0),
        "blue": (0, 0, 1),
        "pink": (1, 0, 0.7),
        "red": (1, 0, 0),
    }

    for ann in annotations:
        if ann["page_num"] != page_num:
            continue
        rect = fitz.Rect(*ann["rect"])
        color = ann["color"]
        if color == "yellow":
            page.add_highlight_annot(rect)
        elif color == "yellow-manual":
            # draw_rect is guaranteed to paint exactly at the given coordinates
            # unlike add_highlight_annot which snaps to nearby text quads
            cx = (rect.x0 + rect.x1) / 2
            cy = (rect.y0 + rect.y1) / 2
            r  = max(rect.width, rect.height) / 2 + 4
            page.draw_circle(fitz.Point(cx, cy), r,
                             color=(1, 0.7, 0), fill=(1, 1, 0), width=1.5)
            page.draw_line(fitz.Point(cx - r, cy), fitz.Point(cx + r, cy),
                           color=(1, 0.4, 0), width=1)
            page.draw_line(fitz.Point(cx, cy - r), fitz.Point(cx, cy + r),
                           color=(1, 0.4, 0), width=1)
        elif color in stroke_map:
            pad = 5
            t = fitz.Rect(rect.x0 - pad, rect.y0 - pad, rect.x1 + pad, rect.y1 + pad)
            a = page.add_rect_annot(t)
            a.set_colors(stroke=stroke_map[color], fill=None)
            a.set_border(width=1)
            a.update()

    occupied = []   # list of (x0, y0, x1, y1) already placed label boxes
    for kabel in kabel_fields:
        if kabel["page_num"] != page_num or not kabel.get("checked", True):
            continue
        text = kabel.get("label", "").strip()
        if not text:
            continue
        r = kabel["rect"]
        fill_rgb = hex_to_fitz_color(kabel.get("label_bg_color", "#ffe066"))
        fs = 9
        pad_h, pad_v = 3, 2
        tw = fitz.get_text_length(text, fontname="helv", fontsize=fs)
        box_w = tw + pad_h * 2
        box_h = fs + pad_v * 2
        lx, ly = _place_label(occupied, float(r[0]), float(r[1]), box_w, box_h)
        occupied.append((lx, ly, lx + box_w, ly + box_h))
        _draw_label_annot(page, lx, ly, box_w, box_h, fill_rgb, text, fs, pad_h)

    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat, alpha=False)
    doc.close()
    return pix.tobytes("png")


# ─────────────────────────────────────────────────────────────────────────────
# Apply / re-number cable labels
# ─────────────────────────────────────────────────────────────────────────────

def apply_labels(kabel_fields, terms):
    cnt_dict = {t.lower(): 0 for t in terms}
    for k in kabel_fields:
        if not k.get("checked", True):
            continue
        term_lower = k.get("term", "").lower()
        cnt = cnt_dict.get(term_lower, 0)
        if k.get("kabel_typ") == "2x RJ45":
            k["label"] = f"{kabel_label_alpha(cnt)}/{kabel_label_alpha(cnt + 1)}"
            cnt += 2
        else:
            k["label"] = kabel_label_alpha(cnt)
            cnt += 1
        cnt_dict[term_lower] = cnt


def _parse_prefix_and_start(pattern):
    """Parse a label pattern like 'Z.01', 'F101', 'AB.001' into (prefix, start_num, width, sep).
    Returns (prefix, sep, start_num, width) or None if unparseable.
    prefix = leading letters, sep = separator (e.g. '.'), start_num = first number, width = zero-pad width."""
    import re as _re
    m = _re.match(r'^([A-Za-z]+)([^0-9]?)(\d+)$', pattern.strip())
    if not m:
        return None
    prefix, sep, num_str = m.group(1), m.group(2), m.group(3)
    return prefix, sep, int(num_str), len(num_str)


def _next_ukv_for_term(kabel_fields, term):
    """Look at ukv_text values of existing entries for `term` and return the next name
    in the detected sequence (e.g. 'Dose 03' → 'Dose 04', 'AB.007' → 'AB.008').
    Returns None when no trailing-number pattern is found."""
    import re as _re
    names = [
        k["ukv_text"] for k in kabel_fields
        if k.get("term", "") == term and k.get("ukv_text", "")
    ]
    # Walk from the last entry backwards; use the first one that ends with digits.
    for name in reversed(names):
        m = _re.match(r'^(.*?)(\d+)$', name.strip())
        if m:
            prefix_part = m.group(1)
            num_str     = m.group(2)
            next_num    = int(num_str) + 1
            width       = len(num_str)
            return f"{prefix_part}{next_num:0{width}d}"
    return None


def apply_labels_with_prefix(kabel_fields, term, pattern):
    """Rename all checked entries of `term` using `pattern` as the first label.
    E.g. pattern='Z.01' → Z.01, Z.02, …; 'F101' → F101, F102, …
    2x-RJ45 entries get two consecutive numbers joined with '/'.
    Unchecked entries are left unchanged."""
    parsed = _parse_prefix_and_start(pattern)
    if not parsed:
        return  # invalid pattern, do nothing
    prefix, sep, start_num, width = parsed
    cnt = start_num
    for k in kabel_fields:
        if k.get("term", "").lower() != term.lower():
            continue
        if not k.get("checked", True):
            continue
        if k.get("kabel_typ") == "2x RJ45":
            l1 = f"{prefix}{sep}{cnt:0{width}d}"
            l2 = f"{prefix}{sep}{cnt+1:0{width}d}"
            k["label"] = f"{l1}/{l2}"
            cnt += 2
        else:
            k["label"] = f"{prefix}{sep}{cnt:0{width}d}"
            cnt += 1


# ─────────────────────────────────────────────────────────────────────────────
# Export helpers
# ─────────────────────────────────────────────────────────────────────────────

def build_excel(kabel_fields):
    wb = Workbook()
    ws = wb.active
    ws.title = "Treffer Export"
    headers = ["Kabelbezeichnung", "Rack", "Bemerkung", "Eingezogen", "Endmontage", "Gemessen und Beschriftet"]
    ws.append(headers)

    def to_excel_color(h):
        h = h.lstrip("#")
        return "FF" + h.upper() if len(h) == 6 else "FFFFFFFF"

    grey = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    light = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    bd = Border(left=Side(style="thin"), right=Side(style="thin"),
                top=Side(style="thin"), bottom=Side(style="thin"))
    al = Alignment(horizontal="center", vertical="center")

    rows = []
    for kabel in kabel_fields:
        if not kabel.get("checked", True):
            continue
        for teil in kabel["label"].split("/"):
            teil = teil.strip()
            if teil:
                rows.append((teil, kabel.get("ukv_text", ""), kabel.get("label_bg_color", "#ffffff")))

    for idx, (text, rack, bg_hex) in enumerate(rows, start=2):
        ws.append([text, rack] + [""] * 4)
        ec = to_excel_color(bg_hex)
        c1 = ws.cell(row=idx, column=1)
        c1.fill = PatternFill(start_color=ec, end_color=ec, fill_type="solid")
        c1.border = bd
        c1.alignment = al
        for col in range(2, 7):
            c = ws.cell(row=idx, column=col)
            c.fill = grey if idx % 2 == 0 else light
            c.border = bd
            c.alignment = al

    for col_cells in ws.columns:
        ml = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = ml + 4

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_annotated_pdf(doc_bytes, kabel_fields, annotations):
    doc = fitz.open(stream=doc_bytes, filetype="pdf")
    stroke_map = {"green": (0, 1, 0), "blue": (0, 0, 1), "pink": (1, 0, 0.7), "red": (1, 0, 0)}

    for ann in annotations:
        pg = doc[ann["page_num"]]
        rect = fitz.Rect(*ann["rect"])
        if ann["color"] == "yellow":
            pg.add_highlight_annot(rect)
        elif ann["color"] in stroke_map:
            pad = 5
            t = fitz.Rect(rect.x0 - pad, rect.y0 - pad, rect.x1 + pad, rect.y1 + pad)
            a = pg.add_rect_annot(t)
            a.set_colors(stroke=stroke_map[ann["color"]], fill=None)
            a.set_border(width=1)
            a.update()

    # Group kabel_fields by page so occupied list is per-page
    from collections import defaultdict
    kabel_by_page = defaultdict(list)
    for kabel in kabel_fields:
        if kabel.get("checked", True):
            kabel_by_page[kabel["page_num"]].append(kabel)

    for pg_num, kabels in kabel_by_page.items():
        pg = doc[pg_num]
        occupied = []
        for kabel in kabels:
            text = kabel["label"].strip()
            if not text:
                continue
            fill_rgb = hex_to_fitz_color(kabel.get("label_bg_color", "#ffe066"))
            r = kabel["rect"]
            fs = 9
            pad_h, pad_v = 3, 2
            tw = fitz.get_text_length(text, fontname="helv", fontsize=fs)
            box_w = tw + pad_h * 2
            box_h = fs + pad_v * 2
            lx, ly = _place_label(occupied, float(r[0]), float(r[1]), box_w, box_h)
            occupied.append((lx, ly, lx + box_w, ly + box_h))
            _draw_label_annot(pg, lx, ly, box_w, box_h, fill_rgb, text, fs, pad_h)

    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Custom CSS
# ─────────────────────────────────────────────────────────────────────────────

st.markdown("""
<style>
    [data-testid="stSidebar"] { min-width: 340px; max-width: 400px; }
    .block-container { padding-top: 1rem; }
    h1 { color: #155d27 !important; }
    .stButton > button { border-radius: 8px; font-weight: 600; }
    .stButton > button[kind="primary"] { background-color: #28a745; border-color: #28a745; }
    .stButton > button[kind="primary"]:hover { background-color: #1e7e34; }
    .cable-row { display: flex; align-items: center; gap: 8px; margin-bottom: 4px; }
    .color-dot { width:12px; height:12px; border-radius:50%; display:inline-block; }
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────────────────────────────────────

with st.sidebar:
    _logo_path = Path(__file__).parent / "logo.svg"
    if _logo_path.exists():
        st.image(str(_logo_path), use_container_width=True)
    else:
        st.markdown("# 🔌 CablingToolPro")
    st.caption("Web-Version")
    st.divider()

    # ── Registration gate ──────────────────────────────────────────────────
    if not st.session_state.user_registered:
        st.markdown("### Zugang")
        st.caption("Bitte Name und E-Mail-Adresse eingeben, um fortzufahren.")
        _reg_name  = st.text_input("Name",    key="_reg_name_input",  placeholder="Vor- und Nachname")
        _reg_email = st.text_input("E-Mail",  key="_reg_email_input", placeholder="name@beispiel.ch")
        if st.button("Weiter →", type="primary", use_container_width=True):
            if _reg_name.strip() and _reg_email.strip() and "@" in _reg_email:
                st.session_state.user_name       = _reg_name.strip()
                st.session_state.user_email      = _reg_email.strip()
                st.session_state.user_registered = True
                st.rerun()
            else:
                st.error("Bitte vollständigen Namen und gültige E-Mail-Adresse eingeben.")
        st.stop()  # block everything below until registered

    # ── PDF Upload ─────────────────────────────────────────────────────────
    uploaded = st.file_uploader("PDF hochladen", type=["pdf"], label_visibility="collapsed")
    st.caption("📂 PDF hochladen")

    if uploaded is not None:
        new_bytes = uploaded.read()
        if new_bytes != st.session_state.doc_bytes:
            st.session_state.doc_bytes = new_bytes
            st.session_state.kabel_fields = []
            st.session_state.annotations = []
            st.session_state.kabel_fields_snap = []
            st.session_state.annotations_snap = []
            st.session_state.pdf_dirty = False
            st.session_state.search_terms = []
            st.session_state.current_page = 0
            st.session_state.export_pdf_bytes = None
            st.session_state.search_ran = False
            _clear_label_widgets()

    if st.session_state.doc_bytes:
        doc_info = fitz.open(stream=st.session_state.doc_bytes, filetype="pdf")
        total_pages = len(doc_info)
        doc_info.close()
        st.success(f"✅ PDF geladen ({total_pages} Seiten)")

    st.divider()
    # ── Projekt laden ──────────────────────────────────────────────────────
    st.caption("📁 Projekt laden (.ctpro)")
    proj_upload = st.file_uploader("Projekt laden", type=["ctpro"], label_visibility="collapsed",
                                   key="proj_uploader")
    if proj_upload is not None:
        try:
            _load_project(proj_upload.read())
            st.rerun()
        except Exception as _e:
            st.error(f"Projekt konnte nicht geladen werden: {_e}")

    st.divider()

    # ── Search ─────────────────────────────────────────────────────────────
    search_input = st.text_input(
        "Suchbegriffe",
        placeholder="z.B. UKV-E00-01, UKV-E00-02",
        help="Mehrere Begriffe mit Komma trennen",
    )

    with st.expander("⚙️ Einstellungen"):
        s2x = st.checkbox("2 x RJ 45 erkennen", value=st.session_state.setting_2x_rj45, key="s2x")
        s2xs = st.checkbox("2x erkennen", value=st.session_state.setting_2x_short, key="s2xs")
        s1x = st.checkbox("RJ 45 erkennen", value=st.session_state.setting_rj45, key="s1x")
        s2xu = st.checkbox("2x UKV erkennen", value=st.session_state.setting_2xukv, key="s2xu")
        st.session_state.setting_2x_rj45 = s2x
        st.session_state.setting_2x_short = s2xs
        st.session_state.setting_rj45 = s1x
        st.session_state.setting_2xukv = s2xu

    do_search = st.button("🔍 Suchen und markieren", use_container_width=True, type="primary")

    if do_search:
        if not st.session_state.doc_bytes:
            st.error("Bitte zuerst ein PDF hochladen!")
        elif not search_input.strip():
            st.warning("Bitte Suchbegriffe eingeben!")
        else:
            terms = [t.strip() for t in search_input.split(",") if t.strip()]
            with st.spinner("Durchsuche PDF…"):
                kf, anns = search_pdf(
                    st.session_state.doc_bytes, terms,
                    st.session_state.setting_2x_rj45,
                    st.session_state.setting_2x_short,
                    st.session_state.setting_rj45,
                    st.session_state.setting_2xukv,
                )
            _clear_label_widgets()
            st.session_state.kabel_fields = kf
            st.session_state.annotations = anns
            st.session_state.kabel_fields_snap = copy.deepcopy(kf)
            st.session_state.annotations_snap = copy.deepcopy(anns)
            st.session_state.pdf_dirty = False
            st.session_state.search_terms = terms
            st.session_state.export_pdf_bytes = None
            st.session_state.search_ran = True
            if kf:
                st.session_state.current_page = kf[0]["page_num"]
            st.rerun()

    # ── Cable list ─────────────────────────────────────────────────────────
    if st.session_state.kabel_fields:
        st.divider()
        st.markdown("**📋 Kabelliste**")

        col_a, col_b = st.columns([1, 1])
        with col_a:
            if st.button("🔄 Umnummerieren", use_container_width=True):
                _clear_component_states()
                apply_labels(st.session_state.kabel_fields, st.session_state.search_terms)
                st.session_state.pdf_dirty = True
                st.rerun()
        with col_b:
            if st.button("\U0001f58a\ufe0f PDF updaten", use_container_width=True, type="primary"):
                st.session_state.kabel_fields_snap = copy.deepcopy(st.session_state.kabel_fields)
                st.session_state.annotations_snap  = list(st.session_state.annotations)
                st.session_state.pdf_dirty = False
                _clear_component_states()
                st.rerun()
        if st.session_state.get("pdf_dirty"):
            st.caption("⚠️ Änderungen noch nicht im PDF – klicke **PDF updaten**")

        if st.button("\U0001f524 Nach Fundstelle sortieren", use_container_width=True):
            _clear_component_states()
            from collections import defaultdict
            term_groups = defaultdict(list)
            for k in st.session_state.kabel_fields:
                term_groups[k.get("term", "")].append(k)
            st.session_state.kabel_fields = [
                k for t in st.session_state.search_terms
                for k in sorted(term_groups[t], key=lambda x: _natural_key(x["ukv_text"]))
            ]
            apply_labels(st.session_state.kabel_fields, st.session_state.search_terms)
            st.session_state.pdf_dirty = True
            st.rerun()

        # Group by search term for display
        terms_order = st.session_state.search_terms
        grouped: dict = {t: [] for t in terms_order}
        for i, k in enumerate(st.session_state.kabel_fields):
            term = k.get("term", "")
            grouped.setdefault(term, []).append((i, k))

        for term in terms_order:
            entries = grouped.get(term, [])
            if not entries:
                continue
            t_idx = terms_order.index(term) if term in terms_order else 0
            current_color = entries[0][1].get("label_bg_color", LABEL_BG_COLORS_HEX[t_idx % len(LABEL_BG_COLORS_HEX)])
            hdr_col, picker_col = st.columns([5, 1])
            with hdr_col:
                st.markdown(
                    f'<span class="color-dot" style="background:{current_color}"></span> '
                    f'**{term}** ({len(entries)} Treffer)',
                    unsafe_allow_html=True,
                )
            with picker_col:
                new_color = st.color_picker("Farbe", current_color,
                                            key=f"termcolor_{term}",
                                            label_visibility="collapsed")
            if new_color != current_color:
                for _, k in entries:
                    k["label_bg_color"] = new_color
                st.session_state.pdf_dirty = True

            # ── Prefix / pattern rename ───────────────────────────────────
            prefix_col, btn_col = st.columns([4, 1])
            with prefix_col:
                prefix_val = st.text_input(
                    "Beschriftungsmuster (z.B. Z.01, F101)",
                    key=f"prefix_{term}",
                    label_visibility="collapsed",
                    placeholder="Muster z.B. Z.01 oder F101 → alle umbenennen",
                )
            with btn_col:
                if st.button("✔", key=f"prefix_btn_{term}", use_container_width=True, type="primary"):
                    if prefix_val.strip():
                        apply_labels_with_prefix(
                            st.session_state.kabel_fields, term, prefix_val.strip()
                        )
                        _clear_component_states()
                        st.session_state.pdf_dirty = True
                        st.rerun()

            # ── Cable list component (edit / delete) ──────────────────────
            # prev_sent_ids: IDs we sent to this component in the PREVIOUS run.
            # A "deletion" is only real if the missing ID was in prev_sent_ids
            # (i.e., the user actually saw it). New entries added since the last
            # render cannot have been deleted by the user — their absence from
            # a stale result is NOT a deletion.
            prev_sent_ids_key = f"cl_prev_sent_ids_{term}"
            prev_sent_ids = list(st.session_state.get(prev_sent_ids_key, []))

            group_items = [
                {
                    "ukv_text": k["ukv_text"],
                    "label": k["label"],
                    "checked": k.get("checked", True),
                    "is_manual": k.get("is_manual", False),
                    "_gi": k.get("_id", i),   # stable ID; falls back to index only for legacy entries
                }
                for i, k in entries
            ]
            orig_ids = [it["_gi"] for it in group_items]

            # Record what we're sending NOW for the next run's reference.
            st.session_state[prev_sent_ids_key] = orig_ids

            result = cable_list_widget(group_items, key=f"cl_{term}")

            if result is not None:
                kf = st.session_state.kabel_fields
                # Build a lookup from stable _id → kabel_fields entry
                id_to_kf = {k.get("_id"): k for k in kf if "_id" in k}

                result_ids = [it["_gi"] for it in result
                              if isinstance(it, dict) and "_gi" in it]
                id_to_res  = {it["_gi"]: it for it in result
                              if isinstance(it, dict) and "_gi" in it}

                # Delete this term's cached value immediately to prevent replay.
                if f"cl_{term}" in st.session_state:
                    del st.session_state[f"cl_{term}"]

                # Discard entirely if any returned ID is unknown or belongs to
                # a different term (stale sendValue after sort/delete/rename).
                if not all(
                    sid in id_to_kf and id_to_kf[sid].get("term", "") == term
                    for sid in id_to_res
                ):
                    pass  # ignore stale result
                else:
                    changed = False
                    for sid, res in id_to_res.items():
                        entry = id_to_kf.get(sid)
                        if entry is None:
                            continue
                        new_label   = res.get("label",   entry["label"])
                        new_checked = res.get("checked", entry.get("checked", True))
                        new_ukv     = res.get("ukv_text", entry["ukv_text"])
                        if (new_label   != entry["label"]
                                or new_checked != entry.get("checked", True)
                                or new_ukv     != entry["ukv_text"]):
                            changed = True
                        entry["label"]    = new_label
                        entry["checked"]  = new_checked
                        entry["ukv_text"] = new_ukv

                    if len(result_ids) < len(orig_ids):
                        # ── Deletion ──────────────────────────────────────
                        # Only IDs that were in prev_sent_ids (component was
                        # actually showing them) can be legitimately deleted.
                        # IDs that appeared AFTER the last render (e.g. just
                        # added via add_position) are NOT deletions.
                        valid_deletable = set(prev_sent_ids) if prev_sent_ids else set(orig_ids)
                        deleted_ids = (set(orig_ids) - set(result_ids)) & valid_deletable
                        if deleted_ids:
                            st.session_state.kabel_fields = [
                                k for k in kf
                                if k.get("_id") not in deleted_ids
                                or k.get("term", "") != term
                            ]
                            _clear_component_states()
                            apply_labels(
                                st.session_state.kabel_fields,
                                st.session_state.search_terms,
                            )
                            changed = True

                    if changed:
                        _clear_component_states()
                        st.session_state.pdf_dirty = True
                        st.rerun()

        # ── Export ─────────────────────────────────────────────────────────
        st.divider()
        active_count = sum(
            len([p for p in k["label"].split("/") if p.strip()])
            for k in st.session_state.kabel_fields
            if k.get("checked", True)
        )
        preis = staffelpreis(active_count)
        st.info(f"**{active_count} Beschriftungen** – CHF {preis:.2f}")

        # Labels are maintained directly in kabel_fields by the component
        def sync_labels():
            pass

        excel_bytes   = build_excel(st.session_state.kabel_fields)
        excel_filename = f"kabelliste_{datetime.date.today()}.xlsx"

        def _on_excel_download():
            send_invoice_email(
                st.session_state.user_name,
                st.session_state.user_email,
                active_count,
                preis,
                excel_bytes,
                excel_filename,
            )

        st.download_button(
            "📥 Excel herunterladen",
            data=excel_bytes,
            file_name=excel_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            on_click=_on_excel_download,
        )

        if st.button("📄 PDF mit Beschriftungen erstellen", use_container_width=True):
            sync_labels()
            with st.spinner("PDF wird erstellt…"):
                pdf_out = build_annotated_pdf(
                    st.session_state.doc_bytes,
                    st.session_state.kabel_fields,
                    st.session_state.annotations,
                )
            st.session_state.export_pdf_bytes = pdf_out
            st.rerun()

        if st.session_state.export_pdf_bytes:
            st.download_button(
                "⬇️ Annotieres PDF herunterladen",
                data=st.session_state.export_pdf_bytes,
                file_name=f"annotiert_{datetime.date.today()}.pdf",
                mime="application/pdf",
                use_container_width=True,
            )

        st.divider()
        st.download_button(
            "💾 Projekt speichern (.ctpro)",
            data=_build_project_zip(),
            file_name=f"projekt_{datetime.date.today()}.ctpro",
            mime="application/octet-stream",
            use_container_width=True,
        )


# ─────────────────────────────────────────────────────────────────────────────
# Main area – PDF viewer
# ─────────────────────────────────────────────────────────────────────────────

if not st.session_state.doc_bytes:
    st.markdown("## 🔌 CablingToolPro – Web Version")
    st.markdown("Suchbegriffs-basiertes Kabelmarkierungstool für PDF-Pläne.")
    st.info("👈 Lade links ein PDF hoch und gib Suchbegriffe ein, um zu starten.")

    st.markdown("### So funktioniert es:")
    st.markdown("""
    1. **PDF hochladen** – Lade deinen Kabelplan als PDF hoch
    2. **Suchbegriffe eingeben** – z. B. `UKV-E00-01, UKV-E00-02` (Komma-getrennt)
    3. **Suchen** – Das Tool findet alle Treffer und erkennt automatisch RJ45-Anschlüsse in der Nähe
    4. **Beschriftungen anpassen** – Labels im Seitenmenü bearbeiten oder umnummerieren
    5. **Exportieren** – Excel-Kabelliste oder annotiertes PDF herunterladen
    """)

else:
    doc_tmp = fitz.open(stream=st.session_state.doc_bytes, filetype="pdf")
    total_pages = len(doc_tmp)
    doc_tmp.close()

    # Page navigation bar
    nav1, nav2, nav3, nav4, nav5 = st.columns([1, 1, 2, 1, 3])
    with nav1:
        if st.button("◀", disabled=st.session_state.current_page == 0):
            st.session_state.current_page -= 1
            st.rerun()
    with nav2:
        if st.button("▶", disabled=st.session_state.current_page >= total_pages - 1):
            st.session_state.current_page += 1
            st.rerun()
    with nav3:
        st.markdown(f"**Seite {st.session_state.current_page + 1} / {total_pages}**")
    with nav4:
        z1, z2, z3 = st.columns([1, 2, 1])
        with z1:
            if st.button("−", disabled=st.session_state.zoom <= 0.5, key="zoom_out"):
                st.session_state.zoom = round(max(0.5, st.session_state.zoom - 0.25), 2)
                st.rerun()
        with z2:
            st.markdown(
                f"<div style='text-align:center;padding-top:6px;font-weight:bold'>{int(st.session_state.zoom * 100)}%</div>",
                unsafe_allow_html=True,
            )
        with z3:
            if st.button("+", disabled=st.session_state.zoom >= 4.0, key="zoom_in"):
                st.session_state.zoom = round(min(4.0, st.session_state.zoom + 0.25), 2)
                st.rerun()
    with nav5:
        if st.session_state.search_ran and st.session_state.kabel_fields:
            pages_with_hits = sorted(set(k["page_num"] for k in st.session_state.kabel_fields))
            if len(pages_with_hits) > 1:
                page_labels = [f"S.{p + 1}" for p in pages_with_hits]
                chosen = st.selectbox(
                    "Springe zu Seite mit Treffern", page_labels, label_visibility="collapsed"
                )
                target = pages_with_hits[page_labels.index(chosen)]
                if target != st.session_state.current_page:
                    st.session_state.current_page = target
                    st.rerun()

    # Sync current label values into kabel_fields before rendering
    for i, k in enumerate(st.session_state.kabel_fields):
        k["label"] = st.session_state.get(f"label_{i}", k["label"])
        k["checked"] = st.session_state.get(f"cb_{i}", k.get("checked", True))

    # Render page from committed snapshot (updated by "PDF updaten" button)
    with st.spinner(""):
        img_bytes = render_page(
            st.session_state.doc_bytes,
            st.session_state.current_page,
            st.session_state.kabel_fields_snap,
            st.session_state.annotations_snap,
            zoom=st.session_state.zoom,
        )

    img_b64 = base64.b64encode(img_bytes).decode()

    # ── Handle right-click "add position" result ──────────────────────────
    viewer_result = pdf_viewer_widget(
        img_b64=img_b64,
        zoom=st.session_state.zoom,
        tx=0,
        ty=0,
        terms=st.session_state.get("search_terms", []),
        key="pdf_viewer",
    )
    # Immediately consume the cached state so it can never be replayed on a
    # later rerun (e.g. triggered by a cable-list component update).
    if "pdf_viewer" in st.session_state:
        del st.session_state["pdf_viewer"]
    if viewer_result and viewer_result.get("action") == "add_position":
        chosen_term = viewer_result["term"]
        # JS computed the PDF coords at right-click time using the formula for
        # /Rotate-90 pages: pdf_x=(nat_h-nat_y)/zoom, pdf_y=nat_x/zoom.
        # We just consume them here and clamp to the page rectangle.
        pdf_x = float(viewer_result.get("pdf_x", 0))
        pdf_y = float(viewer_result.get("pdf_y", 0))

        cur_page = st.session_state.current_page
        _doc = fitz.open(stream=st.session_state.doc_bytes, filetype="pdf")
        _page = _doc.load_page(cur_page)
        _pw = _page.rect.width
        _ph = _page.rect.height
        _doc.close()

        pdf_x = max(0.0, min(pdf_x, _ph))  # after /Rotate-90 swap: x←height axis → clamp to _ph
        pdf_y = max(0.0, min(pdf_y, _pw))  # after /Rotate-90 swap: y←width axis  → clamp to _pw

        # Determine background colour for the chosen term:
        # use the colour already stored on existing entries (may have been changed by user),
        # fall back to the default palette only if no entries exist yet.
        terms_order = st.session_state.get("search_terms", [])
        t_idx = terms_order.index(chosen_term) if chosen_term in terms_order else 0
        kf = st.session_state.kabel_fields
        existing_for_term = [k for k in kf if k.get("term", "") == chosen_term]
        if existing_for_term:
            bg_hex = existing_for_term[0].get(
                "label_bg_color",
                LABEL_BG_COLORS_HEX[t_idx % len(LABEL_BG_COLORS_HEX)],
            )
        else:
            bg_hex = LABEL_BG_COLORS_HEX[t_idx % len(LABEL_BG_COLORS_HEX)]

        # Assign a provisional label (will be corrected when renumbered)
        same_term_items = [k for k in kf if k.get("term", "") == chosen_term and k.get("checked", True)]
        new_label = kabel_label_alpha(len(same_term_items))

        tiny = 8.0  # half-size of the click bounding box in PDF points
        auto_name = _next_ukv_for_term(kf, chosen_term)
        new_entry = {
            "page_num":      cur_page,
            "rect":          [pdf_x - tiny, pdf_y - tiny, pdf_x + tiny, pdf_y + tiny],
            "ukv_text":      auto_name if auto_name else f"Manuell @ ({pdf_x:.0f},{pdf_y:.0f})",
            "term":          chosen_term,
            "label":         new_label,
            "kabel_typ":     None,
            "label_bg_color": bg_hex,
            "checked":       True,
            "is_manual":     True,
            "_id":           _new_id(),
        }

        # Add a crosshair drawing so the position is visible on the PDF
        # Use "yellow-manual" color so render_page uses draw_rect/draw_circle
        # instead of add_highlight_annot (which snaps to text quads)
        st.session_state.annotations.append({
            "page_num": cur_page,
            "rect":     [pdf_x - tiny, pdf_y - tiny, pdf_x + tiny, pdf_y + tiny],
            "color":    "yellow-manual",
        })

        # Insert after the last entry of the same term
        insert_idx = len(kf)  # fallback: append at end
        for idx in range(len(kf) - 1, -1, -1):
            if kf[idx].get("term", "") == chosen_term:
                insert_idx = idx + 1
                break
        kf.insert(insert_idx, new_entry)

        # Renumber all labels so the new entry gets a proper letter/number
        _clear_component_states()
        apply_labels(kf, terms_order)
        # Sync snap immediately so the new marker is visible right away
        st.session_state.kabel_fields_snap = copy.deepcopy(kf)
        st.session_state.annotations_snap  = list(st.session_state.annotations)
        st.session_state.pdf_dirty = False
        st.rerun()

    # Quick-jump buttons below image
    if st.session_state.search_ran and st.session_state.kabel_fields:
        pages_with_hits = sorted(set(k["page_num"] for k in st.session_state.kabel_fields))
        if len(pages_with_hits) > 1:
            st.markdown("**Seiten mit Treffern:**")
            cols = st.columns(min(len(pages_with_hits), 12))
            for ci, pg_num in enumerate(pages_with_hits):
                with cols[ci % 12]:
                    label_str = f"{'→ ' if pg_num == st.session_state.current_page else ''}S.{pg_num + 1}"
                    if st.button(label_str, key=f"jump_{pg_num}"):
                        st.session_state.current_page = pg_num
                        st.rerun()

    # Result summary
    if st.session_state.search_ran:
        if st.session_state.kabel_fields:
            page_hits = [k for k in st.session_state.kabel_fields
                         if k["page_num"] == st.session_state.current_page]
            if page_hits:
                st.success(f"✅ {len(page_hits)} Treffer auf dieser Seite")
        else:
            st.warning("Keine Treffer gefunden.")
